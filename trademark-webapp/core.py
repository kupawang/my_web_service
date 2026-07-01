#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
상표 OA(의견제출통지서) PDF ZIP -> 엑셀 변환 핵심 로직

원본 Claude Skill(build_oa_excel.py)을 웹앱에서 재사용할 수 있도록
- OpenAI 호환 클라이언트 -> Anthropic 공식 API로 교체
- 진행률 콜백(progress_cb) 추가
"""

import os
import re
import io
import json
import base64
import subprocess
import tempfile
import zipfile
import shutil

import fitz  # PyMuPDF
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

import anthropic

MODEL_NAME = "claude-sonnet-4-6"

_ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def sanitize(val):
    if isinstance(val, str):
        return _ILLEGAL.sub('', val)
    return val


# ── ZIP 압축 해제 ──────────────────────────────────────────────────────────

def extract_zip(zip_path: str) -> str:
    tmp_dir = tempfile.mkdtemp(prefix="oa_work_")
    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(tmp_dir)
    return tmp_dir


# ── PDF 텍스트 추출 ────────────────────────────────────────────────────────

def pdf_to_text(pdf_path: str) -> str:
    r = subprocess.run(["pdftotext", "-layout", pdf_path, "-"],
                        capture_output=True, text=True)
    return r.stdout


# ── 상표 이미지 추출 ───────────────────────────────────────────────────────

def get_trademark_images(pdf_path: str, app_no_digits: str = "") -> dict:
    """
    PDF 상표견본 섹션에서 출원상표/선행상표 이미지를 x좌표 기반으로 분리.
    신형 PDF: 라벨(본 출원/인용)이 이미지 위에 있는 구조
    구형 PDF: 라벨이 이전 페이지 하단에, 이미지가 다음 페이지 상단에 있는 구조
    """
    doc = fitz.open(pdf_path)

    label_page = None
    label_page_idx = None
    img_page = None
    img_page_idx = None

    for idx, page in enumerate(doc):
        text = page.get_text()
        if '상표견본이미지' in text or '상표견본 이미지' in text:
            if label_page is None:
                label_page = page
                label_page_idx = idx
            if page.get_images(full=True):
                img_page = page
                img_page_idx = idx

    if label_page is None:
        doc.close()
        return {"app_img": None, "prior_imgs": []}

    if img_page is None and label_page_idx + 1 < len(doc):
        next_page = doc[label_page_idx + 1]
        if next_page.get_images(full=True):
            img_page = next_page
            img_page_idx = label_page_idx + 1

    if img_page is None:
        doc.close()
        return {"app_img": None, "prior_imgs": []}

    target_page = img_page

    words = target_page.get_text('words')
    견본_y = 0
    for w in words:
        if '상표견본' in w[4]:
            견본_y = w[3]
            break

    def collect_labels(word_list, y_min=0):
        _app, _prior = [], []
        j = 0
        while j < len(word_list):
            ww = word_list[j]
            if ww[1] < y_min:
                j += 1
                continue
            if ww[4] == '본' and j + 1 < len(word_list) and word_list[j + 1][4] == '출원':
                _app.append((ww[0] + word_list[j + 1][2]) / 2)
                j += 2
                continue
            if ww[4] == '인용':
                _prior.append((ww[0] + ww[2]) / 2)
            j += 1
        return _app, _prior

    app_label_xs, prior_label_xs = collect_labels(words, y_min=견본_y - 5)

    if not app_label_xs and not prior_label_xs and img_page_idx > 0:
        prev_page = doc[img_page_idx - 1]
        prev_words = prev_page.get_text('words')
        page_height = prev_page.rect.height
        app_label_xs, prior_label_xs = collect_labels(prev_words, y_min=page_height * 0.5)

    xref_data = {}
    for img in target_page.get_images(full=True):
        xref = img[0]
        if xref in xref_data:
            continue
        rects = target_page.get_image_rects(xref)
        if not rects:
            continue
        base_image = doc.extract_image(xref)
        pil_img = Image.open(io.BytesIO(base_image["image"])).convert("RGB")
        xref_data[xref] = {"rects": list(rects), "img": pil_img}

    doc.close()

    all_items = []
    for xref, data in xref_data.items():
        for rect in data["rects"]:
            xc = (rect.x0 + rect.x1) / 2
            yc = (rect.y0 + rect.y1) / 2
            all_items.append((xc, yc, xref, data["img"]))
    all_items.sort(key=lambda t: (round(t[1] / 10), t[0]))

    if not all_items:
        return {"app_img": None, "prior_imgs": []}

    if not app_label_xs and not prior_label_xs:
        return {"app_img": all_items[0][3], "prior_imgs": [t[3] for t in all_items[1:]]}

    def closest(label_x, used):
        best, best_dist = None, float('inf')
        for xc, yc, xref, pimg in all_items:
            key = (xref, round(xc))
            if key in used:
                continue
            dist = abs(xc - label_x)
            if dist < best_dist:
                best_dist, best = dist, (xc, yc, xref, pimg)
        return best

    used = set()
    app_img = None
    if app_label_xs:
        item = closest(app_label_xs[0], used)
        if item:
            app_img = item[3]
            used.add((item[2], round(item[0])))

    prior_imgs = []
    for plx in prior_label_xs:
        item = closest(plx, used)
        if item:
            prior_imgs.append(item[3])
            used.add((item[2], round(item[0])))

    return {"app_img": app_img, "prior_imgs": prior_imgs}


def pil_to_xlsx_image(pil_img, save_path, max_w=100, max_h=70):
    w, h = pil_img.size
    ratio = min(max_w / w, max_h / h, 1.0)
    new_w, new_h = max(1, int(w * ratio)), max(1, int(h * ratio))
    pil_img.resize((new_w, new_h), Image.LANCZOS).save(save_path, "PNG")
    xl_img = XLImage(save_path)
    xl_img.width, xl_img.height = new_w, new_h
    return xl_img, new_w, new_h


# ── AI 호칭 분석 (Anthropic API) ───────────────────────────────────────────

def analyze_trademark_name(pil_img, client: "anthropic.Anthropic") -> dict:
    """
    출원상표 이미지를 Claude API로 분석하여 한국어/영어 호칭 반환.
    반환: {"korean": str, "english": str}
    """
    try:
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        img_b64 = base64.b64encode(buf.getvalue()).decode()

        resp = client.messages.create(
            model=MODEL_NAME,
            max_tokens=200,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": img_b64,
                        },
                    },
                    {
                        "type": "text",
                        "text": (
                            "이 상표 이미지를 보고 한국어 호칭과 영어 호칭을 알려주세요.\n"
                            "- 한국어 호칭: 한글 텍스트가 있으면 그대로, 영문만 있으면 한글 발음으로 표기\n"
                            "- 영어 호칭: 영문 텍스트가 있으면 그대로, 없으면 빈 문자열\n"
                            "- 도형만 있고 문자가 없으면 둘 다 빈 문자열\n"
                            "JSON만 출력하세요. 다른 설명은 절대 붙이지 마세요: "
                            '{"korean": "...", "english": "..."}'
                        ),
                    },
                ],
            }],
        )
        content = ""
        for block in resp.content:
            if getattr(block, "type", None) == "text":
                content += block.text
        m = re.search(r'\{[^}]+\}', content, re.DOTALL)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f"    !! 호칭 분석 오류: {e}")
    return {"korean": "", "english": ""}


# ── OA 요약 추출 ───────────────────────────────────────────────────────────

def extract_oa_summary(text: str) -> str:
    """34조1항7호 거절이유 핵심 내용 추출 (선행상표번호+상표명+지정상품 형식)"""
    sec_m = re.search(r'\[\s*거절이유\s*\d*\s*\]\s*상표법\s*제34조\s*제1항\s*제7호', text)
    if not sec_m:
        sec_m = re.search(r'상표법\s*제34조\s*제1항\s*제7호', text)
    if not sec_m:
        return "(34조1항7호 거절이유 없음)"

    body = text[sec_m.end():]
    end_m = re.search(r'(?:\[\s*거절이유\s*[2-9]|\n\d{4}\.\d{2}\.\d{2}\.\s*\n|끝\.\s*\n)', body)
    if end_m:
        body = body[:end_m.end()]

    body = re.sub(r'이 출원상표는 아래와 같이.*?그러하지 아니합니다\.', '', body, flags=re.DOTALL)
    body = re.sub(r'이 출원상표는 아래와 같이.*?그러하지\s*\n아니합니다\.', '', body, flags=re.DOTALL)
    body = re.sub(r'\d+/\d+\s*\n', '', body)
    body = re.sub(r'끝\.\s*$', '', body, flags=re.MULTILINE)

    lines = [l.strip() for l in body.split('\n') if l.strip()]
    if len(lines) > 25:
        lines = lines[:25] + ["... (이하 생략)"]
    return '\n'.join(lines)


# ── 메타 정보 추출 ─────────────────────────────────────────────────────────

def format_app_no(raw: str) -> str:
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 13 and digits.startswith('40'):
        return f"40-{digits[2:6]}-{digits[6:]}"
    return raw.strip()


def extract_app_no(text: str, filename: str) -> str:
    m = re.search(r'출\s*원\s*번\s*호\s*[:\s]*([\d\-]+)', text)
    if m:
        return format_app_no(m.group(1).strip())
    m2 = re.match(r'(\d+)_', os.path.basename(filename))
    return format_app_no(m2.group(1)) if m2 else ""


def extract_prior_nos(text: str, app_no_digits: str = "") -> list:
    results, seen = [], set()
    img_m = re.search(r'참고\s*:\s*상표견본이미지(.*?)(?:※\s*본\s*통지서|거절이유)', text, re.DOTALL)
    if img_m:
        block = img_m.group(1)
        exclude = {app_no_digits} if app_no_digits else set()
        for m in re.finditer(r'본\s*출\s*원\s*\n?\s*제?([\d]+)호?', block):
            exclude.add(m.group(1))
        for m in re.finditer(r'제([\d]{6,13})호', block):
            num = m.group(1)
            if num not in exclude and num not in seen:
                seen.add(num)
                results.append(f"제{num}호")
        if not results:
            for m in re.finditer(r'\b(4\d{12}|4\d{9}|\d{9,10})\b', block):
                num = m.group(1)
                if num not in exclude and num not in seen:
                    seen.add(num)
                    results.append(f"제{num}호")
    return results


# ── 엑셀 빌드 ──────────────────────────────────────────────────────────────

def build_excel(pdf_files: list, output_path: str, api_key: str,
                 max_count: int = 0, progress_cb=None):
    """
    PDF 파일 목록을 받아 엑셀 생성.
    max_count=0이면 전체 처리.
    progress_cb(done, total, message) 형태의 콜백으로 진행률 보고.
    """
    client = anthropic.Anthropic(api_key=api_key)

    img_dir = tempfile.mkdtemp(prefix="oa_imgs_")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OA 분석"

    hdr_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    cell_font = Font(name="맑은 고딕", size=9)
    name_font = Font(name="맑은 고딕", size=9, bold=True, color="1F4E79")
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
    wrap_left = Alignment(wrap_text=True, vertical="top")
    wrap_bot = Alignment(wrap_text=True, vertical="bottom", horizontal="center")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    target_files = pdf_files if max_count == 0 else pdf_files[:max_count]
    total = len(target_files)

    def report(done, msg):
        if progress_cb:
            progress_cb(done, total, msg)

    records = []
    for i, pdf_path in enumerate(target_files):
        fname = os.path.basename(pdf_path)
        report(i, f"{fname} 처리 중")
        text = pdf_to_text(pdf_path)
        if not text.strip():
            report(i + 1, f"{fname} - 텍스트 없음, 건너뜀")
            continue

        app_no = extract_app_no(text, pdf_path)
        app_no_digits = re.sub(r'\D', '', app_no)
        prior_nos = extract_prior_nos(text, app_no_digits)
        summary = extract_oa_summary(text)

        am = re.search(
            r'출\s*원\s*인\s*성\s*명\s+(.+?)(?:\(\s*특허고객번호|\n\s*주\s*소|\n\s*대\s*리)',
            text, re.DOTALL
        )
        if am:
            applicant = re.sub(r'\s+', ' ', am.group(1)).strip()
            applicant = re.sub(r'\s*\(\s*특허고객번호.*', '', applicant).strip().rstrip(')')
        else:
            applicant = ""

        cm = re.search(r'상\s*품\s*류\s*제\s*(\d+)\s*류', text) or \
            re.search(r'품\s*류\s*제\s*(\d+)\s*류', text)
        trademark_class = cm.group(1) if cm else ""

        dm = re.search(r'발송일자\s*[:\s]*(\d{4}\.\d{2}\.\d{2})', text)
        send_date = dm.group(1) if dm else ""

        imgs = get_trademark_images(pdf_path, app_no_digits)

        name_info = {"korean": "", "english": ""}
        if imgs["app_img"]:
            report(i, f"{fname} - AI 호칭 분석 중")
            name_info = analyze_trademark_name(imgs["app_img"], client)

        records.append({
            "app_no": app_no,
            "applicant": applicant,
            "class": trademark_class,
            "date": send_date,
            "prior_nos": prior_nos,
            "summary": summary,
            "app_img": imgs["app_img"],
            "prior_imgs": imgs["prior_imgs"],
            "korean_name": name_info.get("korean", ""),
            "english_name": name_info.get("english", ""),
        })
        report(i + 1, f"{fname} 완료")

    if not records:
        shutil.rmtree(img_dir, ignore_errors=True)
        raise ValueError("처리된 레코드가 없습니다 (ZIP 안에 유효한 OA PDF가 없는지 확인해주세요).")

    max_prior = max(max(len(r["prior_nos"]), len(r["prior_imgs"])) for r in records)

    COL_APP = 1
    COL_PRIOR_START = 2
    COL_META = COL_PRIOR_START + max_prior
    COL_SUMMARY = COL_META + 1

    headers = {COL_APP: "출원상표", COL_META: "출원정보", COL_SUMMARY: "OA 요약 (34조1항7호)"}
    for i in range(max_prior):
        headers[COL_PRIOR_START + i] = f"선행상표{i + 1}"

    for col, h in headers.items():
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = wrap
        cell.border = border

    for row_idx, rec in enumerate(records, 2):
        ws.row_dimensions[row_idx].height = 90

        korean = rec.get("korean_name", "")
        english = rec.get("english_name", "")
        seen_parts = []
        for p in [korean, english]:
            if p and p not in seen_parts:
                seen_parts.append(p)
        name_text = " / ".join(seen_parts)

        app_cell = ws.cell(row=row_idx, column=COL_APP, value=name_text)
        app_cell.border = border
        app_cell.alignment = wrap_bot
        app_cell.font = name_font

        if rec["app_img"]:
            save_path = os.path.join(img_dir, f"{re.sub(r'\D', '', rec['app_no'])}_app.png")
            xl_img, _, _ = pil_to_xlsx_image(rec["app_img"], save_path, max_w=120, max_h=65)
            xl_img.anchor = f"{get_column_letter(COL_APP)}{row_idx}"
            ws.add_image(xl_img)

        for i in range(max_prior):
            col = COL_PRIOR_START + i
            cell = ws.cell(row=row_idx, column=col, value="")
            cell.border = border
            cell.alignment = wrap
            if i < len(rec["prior_imgs"]) and rec["prior_imgs"][i]:
                save_path = os.path.join(img_dir, f"{re.sub(r'\D', '', rec['app_no'])}_prior{i + 1}.png")
                xl_img, _, _ = pil_to_xlsx_image(rec["prior_imgs"][i], save_path, max_w=90, max_h=55)
                xl_img.anchor = f"{get_column_letter(col)}{row_idx}"
                ws.add_image(xl_img)

        meta_text = (
            f"출원번호: {rec['app_no']}\n"
            f"출원인: {rec['applicant']}\n"
            f"상품류: 제{rec['class']}류\n"
            f"발송일: {rec['date']}"
        )
        meta_cell = ws.cell(row=row_idx, column=COL_META, value=sanitize(meta_text))
        meta_cell.font = cell_font
        meta_cell.alignment = wrap_left
        meta_cell.border = border

        summary_cell = ws.cell(row=row_idx, column=COL_SUMMARY, value=sanitize(rec["summary"]))
        summary_cell.font = cell_font
        summary_cell.alignment = wrap_left
        summary_cell.border = border

    ws.column_dimensions[get_column_letter(COL_APP)].width = 18
    for i in range(max_prior):
        ws.column_dimensions[get_column_letter(COL_PRIOR_START + i)].width = 16
    ws.column_dimensions[get_column_letter(COL_META)].width = 28
    ws.column_dimensions[get_column_letter(COL_SUMMARY)].width = 70
    ws.freeze_panes = "A2"

    wb.save(output_path)
    shutil.rmtree(img_dir, ignore_errors=True)


def run_job(zip_path: str, output_path: str, api_key: str,
            max_count: int = 0, progress_cb=None):
    """ZIP 경로를 받아 압축 해제 -> PDF 목록 -> 엑셀 생성까지 한 번에 처리."""
    oa_dir = extract_zip(zip_path)
    try:
        pdf_files = sorted([
            os.path.join(oa_dir, f)
            for f in os.listdir(oa_dir)
            if f.endswith(".pdf")
        ])
        if not pdf_files:
            raise ValueError("ZIP 안에 PDF 파일이 없습니다.")
        build_excel(pdf_files, output_path, api_key, max_count=max_count, progress_cb=progress_cb)
    finally:
        shutil.rmtree(oa_dir, ignore_errors=True)
